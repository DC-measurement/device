# auther kimto

import paho.mqtt.publish as publish
from time import sleep
import visa
import numpy as np
import time
import openpyxl as px

ID = "USB0::0x2A8D::0x0201::MY57700883::0::INSTR"
r = visa.ResourceManager()
instr = r.get_instrument(ID)

instr.write("SENSe:VOLTage:DC:RANGe{<range>|MIN|}:mV 0.3")
instr.write("CONF:VOLT:DC 10,0.003")
instr.write("TRIG:COUN 10")
instr.write("TRIG:COUN MIN;:SAMP:COUN MIN")
instr.write("CURR:DC:mA 0.1")
instr.write("OUTPUT:STAT ON")


topicV = "34470A/DC/volt"
topicC = "34470A/DC/current"

print('-----------')

# 電圧電流を分けて表示するように改良。（2019/09/29）
# リスト表示を用いた。
# 表示が気に入らない場合はスライス機能を使ってみてもいいかもしれない。

a = ['電圧','電流','抵抗']
print(a)

for V in range(10):

    volts = float(instr.query("MEAS:VOLT:DC? "))
    currents = float(instr.query("MEAS:CURR:DC? "))
    print(volts, "V")
    print(currents, "A")

    # 四則演算機能を利用して抵抗値を算出する機能を付与(2019/09/29)

    x = volts, "V"
    y = currents, "A"
    z = resistance, "Ω"

    # もし、四則演算の過程でオペランドエラーが起こるなら、同じデータ型に変換する箇所があるかもしれない。(2019/09/29)

    resistance = x / y
    print(resistance, "Ω")

    u = [x,y,z]

    print(u)

    #今日の日付・時間を取得（ファイル名に入れるため）
timestr=time.strftime("%Y%m%d-%H%M%S") #年月日時分秒の形式で取得

#新しいExcelの作成とアクティブ化
wb=px.Workbook()
ws=wb.active

#Excelシートの名前を設定
ws.title="result.exe"

#Excel一行目にどんな物理量について計算したか記載するためのリスト
list_target=['電圧','電流','抵抗']

#Excelの1行目に記載していく
for i in range(len(list_target)):
    ws.cell(row=1,column=i+1,value=list_target[i])

#計算結果を保存するためのリスト,今回は5つの物理量A～Cを入れるため3列の配列を作る
Result=np.empty((0,5),int)

#計算部分（ここは計算したい内容によって変える、今回は適当に結果を入れる）
for i in range(10):
    x = volts, "V"
    y = currents, "A"
    z = resistance, "Ω"
    Result=np.append(Result,np.array([[x,y,z]]),axis=0)#結果をさっき作った保存用の配列に入れる

#結果をセルに代入する（※代入は2行目から）
for i in range(len(Result)):
    for j in range(3):
        ws.cell(row=i+2,column=j+1,value=Result[i,j])

#Excelデータをデスクトップに保存する(ユーザー名にお使いのPCの名前を入れてください)
#Excelデータの名前はtest年月日時分秒.xlsx
wb.save(r"/Users/kento/Desktop/test"+timestr+".xlsx")

    
    msg = [{'topic':topicV, 'payload':str(volts)}, ( topicC, str(currents), 0, False)]

    if V == 10:
        break

